from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

def editing_excel(file_name, new_name):
    wb = load_workbook(file_name)
    for s in wb.sheetnames:
        ws = wb[s]
        MAX_ = ws.max_row
        field = (f"A1:X{MAX_}") #TODO: [Maintenablility] X -> is in setting.py VARIABLE:dict_header
        ws.auto_filter.ref = field
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        significance_column = ws['F']
        for cell in significance_column:
            cell.alignment = Alignment(horizontal='center')
    wb.save(new_name)
    wb.close()
