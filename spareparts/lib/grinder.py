import os
import sys

import numpy as np
import pandas as pd
import xlwings as xw
from logzero import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from spareparts.lib.colors import Colors
from spareparts.lib.filters import (
    trash_assemblies,
    trash_description,
    trash_fastener,
    trash_file_name,
    trash_item_number,
    trash_parts_ending_P1_or_A1,
    trash_prp,
    trash_prp1,
    trash_robot,
)
from spareparts.lib.settings import (
    JDEPATH,
    blue,
    dict_header,
    excel_headers,
    headers_bg_hue,
    mauve,
    orange,
    splname,
    temp_jde,
    template1,
    template2,
    tempo_local,
)
from yaspin import Spinner, yaspin

sp = Spinner([
			"[    ]",
			"[=   ]",
			"[==  ]",
			"[=== ]",
			"[ ===]",
			"[  ==]",
			"[   =]",
			"[    ]",
			"[   =]",
			"[  ==]",
			"[ ===]",
			"[====]",
			"[=== ]",
			"[==  ]",
			"[=   ]"
		], 80)


class Spareparts:
    """Generate spareparts list."""

    JDE_TEMP = os.path.join(tempo_local, temp_jde)

    def __init__(self):
        self.jde = self.load_jde_data()
        self.db = pd.DataFrame()
        self.spl = pd.DataFrame()
        self.asm = pd.DataFrame()
        self.elec = pd.DataFrame()
        self.garbage = pd.DataFrame()
        self.nuts = pd.DataFrame()
        self.plates = pd.DataFrame()
        self.gearbox = pd.DataFrame()
        self.drawings = {}

    def generate_spl(self):
        if os.path.exists("SPL.xlsx"):
            raise FileExistsError(
                "Remove or rename the SPL.xlsx in the current folder to able the process to run."
            )
        has_text_reports = os.listdir(".")

        if not has_text_reports:
            raise FileNotFoundError(
                "No text file report has been found in the current folder."
            )

        files = (file for file in Spareparts.listing_txt_files())
        parts = pd.concat(
            [Spareparts.parse_se_report(file) for file in files], ignore_index=True
        )
        self.spl = Spareparts.joining_spl_jde(self.jde, parts)
        self.spl.part_number = (
            self.spl.part_number.str.strip()
        )  # strip part_number column

    def load_db(self):
        """Load the item-level database"""
        db_model = os.path.join(tempo_local, "levels.csv")

        if not os.path.exists(db_model):
            raise FileNotFoundError("No file levels.csv found in user tempo.\n")

        df = pd.read_csv(db_model, dtype={"possibility": str})
        df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
        df.item_number = df.item_number.astype(str)
        df.item_number = df.item_number.str.strip()
        df.possibility = df.possibility.astype(str)
        df.possibility = df.possibility.str.strip()
        self.db = df[["item_number", "possibility"]]
        self.spl = self.spl.join(self.db.set_index("item_number"), on="jdelitm")

    @staticmethod
    def loading_spl(path):
        """load the data from spl list"""
        if not os.path.exists(path):
            raise FileNotFoundError("Check if spl path is correct.")

        spl = pd.read_excel(path, sheet_name="Sheet1")
        spl.columns = spl.columns.str.strip().str.lower().str.replace(" ", "_")
        spl.item_number = spl.item_number.astype("str")
        spl = spl[["item_number"]]
        return spl

    @staticmethod
    def load_jde_data():
        JDE_TEMP = Spareparts.JDE_TEMP
        if os.path.exists(JDE_TEMP):
            answer = input(
                f"Do you want to load the temporary jde? (fast) \n Path:{JDE_TEMP}\n Proceed ([y]/n) ?:"
            )
            if answer.lower() in ["yes", "y"]:
                jde_temp = pd.read_csv(JDE_TEMP)
                return jde_temp
            else:
                print("Process interrupted.")
                sys.exit()
        else:
            with yaspin(sp, side="right", text="Loading the JDE Inventory..."):
                jde_data = Spareparts.extract_jde()
            jde_data.to_csv(JDE_TEMP, index=False)
            return jde_data

    @staticmethod
    def extract_jde():
        """"""
        # add a try - except (in case the file is not found)
        # logger.info()
        df = pd.read_excel(
            JDEPATH,
            sheet_name=0,
            skiprows=[0, 1, 2, 3],
            usecols="A,C,P,E,H,I,K,O,U,X,AA,AR,AT,CB",
            dtype={"Business Unit": int, "Unit Cost": float},
        )
        df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
        df = df[df.business_unit == 101]
        return df

    @staticmethod
    def parse_se_report(fichier):
        """"""
        try:
            # add try and except
            df = pd.read_csv(
                fichier,
                delimiter="\t",
                skiprows=[0, 2],
                header=1,
                names=[
                    "Part Number",
                    "Revision",
                    "DSC_A",
                    "JDELITM",
                    "DIM",
                    "Quantity",
                    "File Name",
                ],
                index_col=False,
                encoding="latin3",
                error_bad_lines=False,
                na_values="-",
            )
        except pd.errors.ParserError as parse_error:
            # Wrong format of text extracted from solidedge.
            logger.error(f" [-][{parse_error}]")
            sys.exit()

        else:
            # clean the columns
            df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
            df["jdelitm"] = df["jdelitm"].str.strip()
            df = Spareparts.replacing_C01(df)
            df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce")
            df = df.groupby(
                ["part_number", "revision", "dsc_a", "dim", "jdelitm", "file_name"],
                as_index=False,
            )["quantity"].sum()
            df = df.replace(r"^-?\s+$", np.nan, regex=True)
            df = df.dropna(subset=["part_number", "jdelitm"])
            # give the module number
            module_number = os.path.splitext(os.path.basename(fichier))[0]
            df["module"] = module_number
            logger.info(" [+][\t %s }\t]" % module_number)
            return df

        finally:
            df = None

    @staticmethod
    def listing_txt_files():
        """"""
        return (file for file in os.listdir(".") if file.endswith(".txt"))

    @staticmethod
    def replacing_C01(df):
        """Replacing 123456_C01 to 123456."""
        pat = r"(?P<number>\d{6})(?P<suffixe>_C\d{2})"
        repl = lambda m: m.group("number")
        df["part_number"] = df["part_number"].str.replace(pat, repl)
        return df

    @staticmethod
    def joining_spl_jde(jde, parts):
        """transform the jde column to string format
        join the parts documents with the jde on jdelitm column
        and sort it on column:module
        """
        jde.item_number = jde.item_number.astype(str)
        spl = parts.join(jde.set_index("item_number"), on="jdelitm").sort_values(
            "module"
        )
        return spl

    def part_type(self):
        """create a column type --> .par .psm .asm"""
        self.spl["type"] = self.spl.file_name.str.split(".").str[-1].str.strip()
        self.spl.type = self.spl.type.str.lower()

    def lines_numbers(self):
        logger.info(
            "\n\n"
            "Qty/Groups :\n"
            "-------------------------\n"
            f"spl       :\t{self.spl.shape[0]}\n"
            f"garbage   :\t{self.garbage.shape[0]}\n"
            f"plates    :\t{self.plates.shape[0]}\n"
            f"elec      :\t{self.elec.shape[0]}\n"
            f"asm       :\t{self.asm.shape[0]}\n"
            f"nuts      :\t{self.nuts.shape[0]}\n"
            "-------------------------\n\n"
        )

    @yaspin(sp, side="right", text="Creating excel file, do not close the window ")
    def create_excel(self, given_name_xlsx):
        """fill the tabs in excel file with the dataframes"""
        tabs = {
            "nuts": self.nuts,
            "asm": self.asm,
            "plates": self.plates,
            "elec": self.elec,
            "gearbox": self.gearbox,
            "garbage": self.garbage,
            "spl": self.spl,
        }
        wb = xw.Book()  # this will create a new workbook
        for tab in tabs.keys():
            sht = wb.sheets.add(tab)
        for tab, df in tabs.items():
            sht = wb.sheets[
                tab
            ]  # skip the Sheet1 and create spl within a loop for all tab
            sht.range("A1").value = excel_headers  # insert headers (horizontal)
            sht.range("A1:R1").api.Font.Bold = True  # bold headers (horizontal)
            for rang, color in headers_bg_hue.items():
                xw.Range(rang).color = color
            for colum, data in dict_header.items():
                sht.range(colum).options(index=False, header=False).value = df[data]
            sht.autofit()
        wb.sheets[-1].delete()
        wb.save(given_name_xlsx)
        wb.close()
        logger.info(f"{template1}: created")

    @staticmethod
    @yaspin(sp, side="right", text="Editing excel file, do not close the window ")
    def edit_excel(file_name, new_name):
        wb = load_workbook(file_name)
        for s in wb.sheetnames:
            ws = wb[s]
            MAX_ = ws.max_row
            field = f"A1:X{MAX_}"
            ws.auto_filter.ref = field
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            significance_column = ws["F"]
            for cell in significance_column:
                cell.alignment = Alignment(horizontal="center")
        wb.save(new_name)
        wb.close()
        logger.info(f"{template2}: created")

    def refine(self):
        ambiguous = self.spl[
            ~(
                (self.spl.possibility == "1")
                | (self.spl.possibility == "2")
                | (self.spl.possibility == "3")
            )
        ]
        ambiguous_items = (
            ambiguous.part_number.str.strip().tolist()
        )  # Whitespaces stripped here
        for itm in ambiguous_items:
            mdl = self.spl.loc[itm, "module"]  # module => mdl
            self.spl.possibility[
                self.spl.part_number == itm, "possibility"
            ] = self.db.loc[itm, mdl]

    @Colors.obsolete(mauve)
    @Colors.meter_foot(blue)
    @Colors.electric(["Electric Component"], orange)
    def extraction(splname, workbook, sht_name):
        df = pd.read_excel(splname, sheet_name=sht_name)
        sht = workbook.sheets[sht_name]
        return (df, sht)

    extraction = staticmethod(extraction)

    @staticmethod
    def add_colors(selected_file, sheet_spl):
        wb = xw.Book(selected_file)
        Spareparts.extraction(selected_file, wb, sheet_spl)
        return wb

    @yaspin(sp, side="right", text="Editing excel file, do not close the window ")
    def colors_excel(self, selected_file, new_file):
        args = {
            "nuts": self.nuts,
            "asm": self.asm,
            "plates": self.plates,
            "elec": self.elec,
            "garbage": self.garbage,
            "spl": self.spl,
        }
        for tab in args:
            wb = Spareparts.add_colors(selected_file, tab)
        wb.save(new_file)
        wb.close()
        logger.info(f"{splname}: created")

    @staticmethod
    def log_report(_df, df_name):
        if _df.shape[0] == 0:
            pass
        else:
            _df["groupe"] = df_name
            logger.info("\n"+ _df[["groupe", "part_number", "description_1"]].to_string())

    def strain(self):
        """Filters of unwanted parts here."""

        # --------------------------------------------------------------------#
        #                                                                    #
        #                           START FILTERS HERE                       #
        #                                                                    #
        # --------------------------------------------------------------------#

        # === plates ===
        plates_prp1 = ["Aluminium", "Stainless Steel", "Steel"]
        self.spl, self.garbage, _plates = trash_prp1(
            self.spl, self.garbage, prp1=plates_prp1
        )
        self.garbage = pd.concat([self.garbage, _plates]).drop_duplicates(keep=False)
        Spareparts.log_report(_plates, "_plates")

        # === fasteners ===
        self.spl, self.garbage, _nuts = trash_fastener(self.spl, self.garbage)
        self.garbage = pd.concat([self.garbage, _nuts]).drop_duplicates(keep=False)
        Spareparts.log_report(_nuts, "_nuts")

        # === assemblies ===
        self.spl, self.garbage, _asm = trash_assemblies(
            self.spl, self.garbage
        )  # _ASM SEEMS TO BE THE EXCEPTION
        self.garbage = pd.concat([self.garbage, _asm]).drop_duplicates(keep=False)
        Spareparts.log_report(_asm, "_asm")

        # === uncategorized ===
        self.garbage_prp1 = [
            "Sign & Label",
            "Plumbing Hardware",
            "Pièce Manufacturée Magasin",
        ]
        self.spl, self.garbage, _uncatego = trash_prp1(
            self.spl, self.garbage, prp1=self.garbage_prp1
        )
        self.garbage = pd.concat([self.garbage, _uncatego]).drop_duplicates(keep=False)
        Spareparts.log_report(_uncatego, "_uncatego")

        # === robot ===
        self.spl, self.garbage, _robot = trash_robot(self.spl, self.garbage)
        self.garbage = pd.concat([self.garbage, _robot]).drop_duplicates(keep=False)
        Spareparts.log_report(_robot, "_robot")

        # === gripper ===

        contents_of_gripper = [
            "PT1124830",
            "PT0078604",
            "PT0078603",
            "24104091",
            "24101598",
            "24101597",
            "171257",
            "171259",
            "171256",
            "171255",
            "24100056",
            "PT0078602",
            "PT0078601",
            "EEG58C7007P-1",
            "24100360",
            "EEG58C6002P-6",
            "54010220",
            "24300030",
            "24104854",
            "24104591",
            "24104548",
            "162925_EEG58C",
            "171228",
        ]
        self.spl, self.garbage, _inside_gripper = trash_item_number(
            self.spl, self.garbage, list_parts=contents_of_gripper
        )
        self.garbage = pd.concat([self.garbage, _inside_gripper]).drop_duplicates(
            keep=False
        )
        Spareparts.log_report(_inside_gripper, "_inside_gripper")

        # === industrial ===
        self.spl, self.garbage, _industrial = trash_prp(
            self.spl, self.garbage, prp1=["Industrial Engine"], prp2=["Engine Parts"]
        )
        self.garbage = pd.concat([self.garbage, _industrial]).drop_duplicates(
            keep=False
        )
        Spareparts.log_report(_industrial, "_industrial")

        # === _furniture ===
        self.spl, self.garbage, _furniture = trash_prp(
            self.spl, self.garbage, prp1=["Factory Furniture"], prp2=["Tape"]
        )
        self.garbage = pd.concat([self.garbage, _furniture]).drop_duplicates(keep=False)
        Spareparts.log_report(_furniture, "_furniture")

        # === _gearbox ===
        self.spl, self.garbage, _gearbox = trash_prp(
            self.spl,
            self.garbage,
            prp1=["Mechanical Component"],
            prp2=["Gearbox, Gear, Rack & Pinion", "Gear Motor & Motor"],
        )
        self.garbage = pd.concat([self.garbage, _gearbox]).drop_duplicates(keep=False)
        Spareparts.log_report(_gearbox, "_gearbox")

        # === _grommet ===
        self.spl, self.garbage, _grommet = trash_description(
            self.spl, self.garbage, keyword="GROMMET;RUBBER"
        )
        self.garbage = pd.concat([self.garbage, _grommet]).drop_duplicates(keep=False)
        Spareparts.log_report(_grommet, "_grommet")

        # === _pneu_frl ===
        self.spl, self.garbage, _pneu_frl = trash_description(
            self.spl,
            self.garbage,
            keyword=r"PNEU\.F\.R\.L",
            description="description_1",
        )
        self.garbage = pd.concat([self.garbage, _pneu_frl]).drop_duplicates(keep=False)
        Spareparts.log_report(_pneu_frl, "_pneu_frl")

        # === _clamp ===
        self.spl, self.garbage, _clamp = trash_description(
            self.spl,
            self.garbage,
            keyword="CLAMP;TRANSPORT UNIT",
            description="description_2",
        )
        self.garbage = pd.concat([self.garbage, _clamp]).drop_duplicates(keep=False)
        Spareparts.log_report(_clamp, "_clamp")

        # === electric ===
        self.spl, self.garbage, _elec = trash_prp(
            self.spl,
            self.garbage,
            prp1=["Electric Component"],
            prp2=[
                "Cable Tray & Cable Carrier",
                "Conduits & fittings",
                "Enclosures",
                "Sensors",
                "Lights & bulbs",
                "Switches",
                "General hardware",
                "Stickers",
                "Buttons & pilot lights",
                "Connectors & crimps",
            ],
        )
        self.garbage = pd.concat([self.garbage, _elec]).drop_duplicates(keep=False)
        # logger.info(f"***_elec: {_elec}")

        Spareparts.log_report(_elec, "_elec")

        # === _par ===
        self.spl, self.garbage, _par = trash_file_name(
            self.spl, self.garbage, keyword=r"^par\s*$"
        )
        self.garbage = pd.concat([self.garbage, _par]).drop_duplicates(keep=False)
        Spareparts.log_report(_par, "_par")

        # === _P1_A1 ===
        self.spl, self.garbage, _P1_A1 = trash_parts_ending_P1_or_A1(
            self.spl, self.garbage
        )
        self.garbage = pd.concat([self.garbage, _P1_A1]).drop_duplicates(keep=False)
        Spareparts.log_report(_P1_A1, "_P1_A1")

        # === _collar ===
        collar = r"COLLAR"
        self.spl, self.garbage, _collar = trash_description(
            self.spl, self.garbage, keyword=collar
        )
        self.garbage = pd.concat([self.garbage, _collar]).drop_duplicates(keep=False)
        Spareparts.log_report(_collar, "_collar")

        # === fpmr ===
        fpmr = r"FIXING PLATE MOTOR ROLLER"
        self.spl, self.garbage, _fpmr = trash_description(
            self.spl, self.garbage, keyword=fpmr
        )
        self.garbage = pd.concat([self.garbage, _fpmr]).drop_duplicates(keep=False)
        Spareparts.log_report(_fpmr, "_fpmr")

        # === _fit ===
        fit = r"PNEU.FIT.NIPPLE"
        self.spl, self.garbage, _fit = trash_description(
            self.spl, self.garbage, keyword=fit
        )
        self.garbage = pd.concat([self.garbage, _fit]).drop_duplicates(keep=False)
        Spareparts.log_report(_fit, "_fit")

        # === miscellaneous ===
        rejected_parts = [
            "DPP95A1530S-3",
            "136918",
            "216081",
            "162463",
            "146660",
            "02479",
            "EEG59F2164P-9151967",
        ]
        self.spl, self.garbage, _misc = trash_item_number(
            self.spl, self.garbage, rejected_parts
        )
        self.garbage = pd.concat([self.garbage, _misc]).drop_duplicates(keep=False)
        Spareparts.log_report(_misc, "_misc")

        # === _keysquare ===
        self.spl, self.garbage, _keysquare = trash_description(
            self.spl, self.garbage, keyword="KEY;SQUARE"
        )
        self.garbage = pd.concat([self.garbage, _keysquare]).drop_duplicates(keep=False)
        Spareparts.log_report(_keysquare, "_keysquare")

        # === _stud ===
        self.spl, self.garbage, _stud = trash_description(
            self.spl, self.garbage, keyword="STUD"
        )
        self.garbage = pd.concat([self.garbage, _stud]).drop_duplicates(keep=False)
        Spareparts.log_report(_stud, "_stud")

        # === _suppliers ===
        self.spl, self.garbage, _suppliers = trash_description(
            self.spl, self.garbage, keyword="SOUS-TRAITANCE RECOUVREMENT"
        )
        self.garbage = pd.concat([self.garbage, _suppliers]).drop_duplicates(keep=False)
        Spareparts.log_report(_suppliers, "_suppliers")

        # === weldnut ===
        self.spl, self.garbage, _weldnut = trash_description(
            self.spl, self.garbage, keyword="WELD NUT"
        )
        self.garbage = pd.concat([self.garbage, _weldnut]).drop_duplicates(keep=False)
        Spareparts.log_report(_weldnut, "_weldnut")

        # === _transparent ===
        self.spl, self.garbage, _transparent = trash_prp(
            self.spl, self.garbage, prp1=["POLYCARBONATE PLATE"], prp2=["TRANSPARENT"]
        )
        self.garbage = pd.concat([self.garbage, _transparent]).drop_duplicates(
            keep=False
        )
        Spareparts.log_report(_transparent, "_transparent")

        # === _hoffman_pkg ===
        self.spl, self.garbage, _hoffman_pkg = trash_item_number(
            self.spl, self.garbage, list_parts=["PT1173377", "PT1173378"]
        )
        self.garbage = pd.concat([self.garbage, _hoffman_pkg]).drop_duplicates(
            keep=False
        )
        Spareparts.log_report(_hoffman_pkg, "_hoffman_pkg")

        # === EYE BOLT;SHOULDER ===
        self.spl, self.garbage, _eye = trash_description(
            self.spl, self.garbage, keyword="EYE BOLT;SHOULDER"
        )
        self.garbage = pd.concat([self.garbage, _eye], sort=True).drop_duplicates(
            keep=False
        )
        Spareparts.log_report(_eye, "_eye")

        # === POLYCARBONATE PLATE ===
        self.spl, self.garbage, _polycarbonate = trash_description(
            self.spl, self.garbage, keyword="POLYCARBONATE PLATE"
        )
        self.garbage = pd.concat(
            [self.garbage, _polycarbonate], sort=True
        ).drop_duplicates(keep=False)
        Spareparts.log_report(_polycarbonate, "_polycarbonate")

        # --------------------------------------------------------------------#
        #                           END FILTERS HERE                         #
        # --------------------------------------------------------------------#

        self.spl = self.spl.drop_duplicates(keep=False)
        self.asm = _asm
        self.elec = _elec
        self.nuts = _nuts
        self.plates = _plates
        self.gearbox = _gearbox
        # Garbage include all filtered parts
        self.garbage = pd.concat(
            [
                self.garbage,
                _uncatego,
                _inside_gripper,
                _industrial,
                _furniture,
                _grommet,
                # _manif,
                _pneu_frl,
                _clamp,
                _par,
                _collar,
                _fpmr,
                _fit,
                _misc,
                _stud,
                _keysquare,
                _stud,
                _suppliers,
                _weldnut,
                _transparent,
                _hoffman_pkg,
                _eye,
                _polycarbonate,
            ],
            ignore_index=True,
            sort=False,
        ).drop_duplicates(keep=False)

    def equivalences(self):
        """return dictionnary > {ptnumber:drawing number}"""
        prt_num = self.spl.part_number
        prt_num = prt_num[
            prt_num.str.contains(
                r"\bPT\d{7}\b|\bEEG.*\b|\bEPT.*\b", na=False, regex=True
            )
        ]  # debug r'PT\d{7}|EEG|EPT'
        prt_num = prt_num.str.strip().tolist()
        # drawing number is found in JDE:'drawing_number'
        _drawing = "drawing_number"
        _item = "item_number"
        itm_num_drawing = self.jde[[_item, _drawing]]
        itm_num_drawing = itm_num_drawing.set_index(_drawing)
        equivalences = {}
        for i in prt_num:
            if i in itm_num_drawing.index.tolist():
                if len(itm_num_drawing.loc[i, _item]) == 6:
                    equivalences[i] = itm_num_drawing.loc[i, _item]
        self.drawings = equivalences

    def drawing_number(self):
        """add column """
        self.spl["number"] = self.spl["part_number"].map(
            self.drawings, na_action="ignore"
        )

    @staticmethod
    def del_templates():
        logger.info(f"removing {template1}, {template2}")
        os.remove(template1)
        os.remove(template2)
        logger.info(f"{template1} - {template2}: deleted")

    @staticmethod
    def prompt_confirmation():
        "ask user to resume the program"
        logger.info(f"Run: {__file__}")
        answer = input("Proceed ([y]/n) ?:  ")
        if answer.lower() in ["yes", "y"]:
            pass
        else:
            print("Process has stopped.")
            sys.exit()
