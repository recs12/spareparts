from openpyxl import load_workbook

def autofilter(file_name):
    wb = load_workbook(file_name)
    for s in wb.sheetnames:
        ws = wb[s]
        MAX_ = ws.max_row
        field = (f"A1:T{MAX_}")
        ws.auto_filter.ref = field
    wb.save("auto_with_filters.xlsx")
    print(f"excel file created: auto_with_filters.xlsx")

